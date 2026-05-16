"""Export endpoints: HTML / XLSX for executions, XLSX for project cases."""
import base64
import html as html_lib
import io
import os
from collections import defaultdict
from datetime import datetime, timezone
from urllib.parse import urlparse


def _attachment_data_url(
    file_path: str | None, content_type: str | None, limit_bytes: int,
) -> str | None:
    if not file_path or not os.path.exists(file_path):
        return None
    if os.path.getsize(file_path) > limit_bytes:
        return None
    with open(file_path, "rb") as handle:
        encoded = base64.b64encode(handle.read()).decode("ascii")
    mime = content_type or "application/octet-stream"
    return f"data:{mime};base64,{encoded}"

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from fastapi import Depends, HTTPException, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, selectinload

from testjam.auth.dependencies import get_current_user, require_project_access
from testjam.core.config import settings
from testjam.database import get_db
from testjam.models.execution import TestExecution, TestResult
from testjam.models.testcase import TestCase, TestSuite
from testjam.models.user import User
from testjam.routers.executions import executions_router, projects_router
from testjam.routers.executions._helpers import load_execution_full
from testjam.services.settings import get_settings as get_app_settings
from testjam.services.timezones import format_in_user_zone, user_zone, user_zone_label


@executions_router.get("/{id}/export/xlsx")
def export_execution_xlsx(
    id: int,
    db: Session = Depends(get_db),
    current: User = Depends(get_current_user),
):
    ex = load_execution_full(db, id)
    if not ex:
        raise HTTPException(status_code=404, detail="Not found")

    wb = openpyxl.Workbook()
    tz_label = user_zone_label(current)

    # Summary sheet
    ws_sum = wb.active
    ws_sum.title = "Summary"
    launched_by = ex.token_name or (ex.created_by.username if ex.created_by else ex.triggered_by or "")
    ws_sum.append(["Title", ex.title])
    ws_sum.append(["Status", ex.status])
    ws_sum.append(["Type", ex.type])
    ws_sum.append(["Environment", ex.environment or ""])
    ws_sum.append(["Version", ex.project_version.name if ex.project_version else ""])
    ws_sum.append(["Launched by", launched_by])
    ws_sum.append([f"Started at ({tz_label})", format_in_user_zone(ex.started_at, current, "%Y-%m-%d %H:%M:%S")])
    ws_sum.append([f"Finished at ({tz_label})", format_in_user_zone(ex.finished_at, current, "%Y-%m-%d %H:%M:%S")])
    ws_sum.append([
        "Generated",
        f"{format_in_user_zone(datetime.now(timezone.utc), current, '%Y-%m-%d %H:%M %Z')} by {current.username}",
    ])

    # Results sheet
    ws_res = wb.create_sheet("Results")
    ws_res.append(["Test Case", "Status", "Executed by", f"Executed at ({tz_label})", "Duration (ms)", "Comment"])
    for r in ex.results:
        tc_name = r.test_case.name if r.test_case else str(r.test_case_id)
        ws_res.append([
            tc_name,
            r.status,
            r.executed_by or "",
            format_in_user_zone(r.executed_at, current, "%Y-%m-%d %H:%M:%S"),
            r.duration_ms or "",
            r.comment or "",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"execution_{id}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@executions_router.get("/{id}/export/html")
def export_execution_html(
    id: int,
    request: Request,
    db: Session = Depends(get_db),
    current: User = Depends(get_current_user),
):
    ex = (
        db.query(TestExecution)
        .options(
            selectinload(TestExecution.results)
            .selectinload(TestResult.test_case)
            .selectinload(TestCase.suite),
            selectinload(TestExecution.results)
            .selectinload(TestResult.test_case)
            .selectinload(TestCase.steps),
            selectinload(TestExecution.results).selectinload(TestResult.step_results),
            selectinload(TestExecution.results).selectinload(TestResult.attachments),
            selectinload(TestExecution.attachments),
        )
        .filter(TestExecution.id == id)
        .first()
    )
    if not ex:
        raise HTTPException(status_code=404, detail="Not found")

    # Bulk-load all suites of the project so the ancestor walk uses an in-memory
    # parent_suite_id lookup instead of `s.parent` lazy-loading per ancestor.
    project_suites: dict[int, TestSuite] = {}
    if ex.project_id:
        for s in db.query(TestSuite).filter(TestSuite.project_id == ex.project_id).all():
            project_suites[s.id] = s

    def e(s): return html_lib.escape(str(s)) if s is not None else ""

    user_tz = user_zone(current)

    def fmt_date(dt):
        if not dt:
            return "—"
        local = dt.astimezone(user_tz)
        ms = local.microsecond // 1000
        return local.strftime("%-d %b %Y, %H:%M:%S") + f".{ms:03d}"

    def fmt_dur(ms):
        if ms is None: return ""
        return f"{ms}ms" if ms < 1000 else f"{ms/1000:.1f}s"

    app_settings = get_app_settings(db)
    # Prefer the caller's origin (Referer) so links resolve from wherever the
    # browser invoked Export HTML; fall back to configured site_url, then to
    # the backend's request URL.
    base_url = ""
    referer = request.headers.get("referer", "")
    if referer:
        try:
            p = urlparse(referer)
            if p.scheme and p.netloc:
                base_url = f"{p.scheme}://{p.netloc}"
        except Exception:
            pass
    if not base_url and app_settings.site_url:
        base_url = app_settings.site_url.rstrip("/")
    if not base_url:
        base_url = f"{request.url.scheme}://{request.url.netloc}"
    exec_url = f"{base_url}/executions/{id}/run"

    launched_by = ex.token_name or (ex.created_by.username if ex.created_by else ex.triggered_by or "—")
    counts = {"passed": 0, "failed": 0, "blocked": 0, "not_run": 0}
    for r in ex.results:
        counts[r.status] = counts.get(r.status, 0) + 1
    total = sum(counts.values())

    # Build compact header meta line
    meta_items = []
    project_name = ex.project.name if ex.project else ""
    if project_name:
        meta_items.append(e(project_name))
    if ex.environment:
        meta_items.append(e(ex.environment))
    if ex.project_version:
        meta_items.append(f"v{e(ex.project_version.name)}")
    if launched_by and launched_by != "—":
        meta_items.append(f"by {e(launched_by)}")
    dates_str = fmt_date(ex.started_at)
    if ex.finished_at:
        dates_str += f" → {fmt_date(ex.finished_at)}"
    if dates_str and dates_str != "—":
        meta_items.append(dates_str)
    meta_html = "".join(f'<span>{m}</span>' for m in meta_items)

    # Header extras: execution URL + attachments
    hextras = []
    if exec_url:
        hextras.append(f'<a class="hlink" href="{e(exec_url)}" target="_blank">↗ View in {e(app_settings.app_name)}</a>')
    atts = ex.attachments or []
    if atts:
        att_links = []
        inline_limit_bytes = app_settings.export_inline_attachment_mb * 1024 * 1024
        for a in atts:
            data_url = _attachment_data_url(a.file_path, a.content_type, inline_limit_bytes)
            if data_url is None:
                att_links.append(
                    f'<span class="hatt-missing" title="Attachment exceeds inline limit or missing">{e(a.filename)}</span>',
                )
                continue
            att_links.append(
                f'<a href="{data_url}" download="{e(a.filename)}" target="_blank" rel="noopener">{e(a.filename)}</a>',
            )
        hextras.append(f'<span class="hatts">Attachments: {" ".join(att_links)}</span>')
    hextras_html = f'<div class="hextras">{" ".join(hextras)}</div>' if hextras else ""

    # Group results by suite_id; collect ancestor suites from in-memory map.
    results_by_suite_id: dict[int | None, list] = defaultdict(list)
    suites_by_id: dict[int, TestSuite] = {}

    for r in ex.results:
        suite = r.test_case.suite if r.test_case else None
        suite_id = suite.id if suite else None
        results_by_suite_id[suite_id].append(r)
        s = suite
        while s is not None:
            if s.id not in suites_by_id:
                suites_by_id[s.id] = s
            s = project_suites.get(s.parent_suite_id) if s.parent_suite_id else None

    # parent_id → [child suite ids]
    children_of: dict[int | None, list[int]] = defaultdict(list)
    for sid, suite in suites_by_id.items():
        parent_id = suite.parent_suite_id if suite.parent_suite_id in suites_by_id else None
        children_of[parent_id].append(sid)

    def _root_id(suite_id: int | None) -> int | None:
        if suite_id is None or suite_id not in suites_by_id:
            return None
        pid = suites_by_id[suite_id].parent_suite_id
        return suite_id if (pid is None or pid not in suites_by_id) else _root_id(pid)

    # Preserve root order from ex.results insertion order
    root_order: list[int | None] = []
    seen_roots: set[int | None] = set()
    for r in ex.results:
        suite = r.test_case.suite if r.test_case else None
        rid = _root_id(suite.id if suite else None)
        if rid not in seen_roots:
            seen_roots.add(rid)
            root_order.append(rid)

    def render_steps(r) -> str:
        steps = sorted(r.test_case.steps, key=lambda s: s.order) if r.test_case and r.test_case.steps else []
        sr_by_step = {sr.step_id: sr for sr in r.step_results}
        if not steps:
            return ""
        parts = []
        for step in steps:
            sr = sr_by_step.get(step.id)
            st = sr.status if sr else "not_run"
            is_step_error = st in ("failed", "blocked")
            log = sr.log_output if sr else None
            comment = sr.comment if sr else None
            dur = fmt_dur(sr.duration_ms) if sr else ""
            step_meta_parts = [p for p in [dur, fmt_date(r.executed_at)] if p and p != "—"]
            step_meta = " · ".join(step_meta_parts)
            has_body = step.expected_result or comment or log
            open_attr = 'open data-error="1"' if is_step_error else ""
            expected_html = f'<div class="sexpected"><span class="slabel">Expected</span>{e(step.expected_result)}</div>' if step.expected_result else ""
            comment_html = f'<div class="scomment">💬 {e(comment)}</div>' if comment else ""
            log_html = f'<pre class="slog">{e(log)}</pre>' if log else ""
            body_html = f'<div class="sbody">{expected_html}{comment_html}{log_html}</div>' if has_body else ""
            parts.append(f"""<details class="step {e(st)}" {open_attr}>
        <summary class="step-hd">
          <span class="schev">▶</span>
          <span class="snum">{step.order}.</span>
          <span class="kpill kp-{e(st)}">{e(step.step_type.capitalize())}</span>
          <span class="saction">{e(step.action)}</span>
          {f'<span class="tmeta">{e(step_meta)}</span>' if step_meta else ""}
        </summary>
        {body_html}
      </details>""")
        return "\n      ".join(parts)

    def render_result(r) -> str:
        st = r.status
        is_error = st in ("failed", "blocked")
        dur = fmt_dur(r.duration_ms)
        meta_parts = [p for p in [e(r.executed_by), dur, fmt_date(r.executed_at)] if p and p != "—"]
        meta = " · ".join(meta_parts) if meta_parts else ""
        steps_html = render_steps(r)
        comment_html = f'<div class="comment">💬 {e(r.comment)}</div>' if r.comment else ""
        steps_div = f'<div class="steps">{steps_html}</div>' if steps_html else ""
        body = f'<div class="tbody">{comment_html}{steps_div}</div>' if (comment_html or steps_html) else ""
        open_attr = 'open data-error="1"' if is_error else ""
        return f"""<details class="test" {open_attr}>
      <summary class="thd">
        <span class="chevron">▶</span>
        <span class="kpill kp-{e(st)}">Test</span>
        <span class="tname">{e(r.test_case.name if r.test_case else str(r.test_case_id))}</span>
        {f'<span class="tmeta">{e(meta)}</span>' if meta else ""}
      </summary>
      {body}
    </details>"""

    def all_results_for(suite_id: int) -> list:
        out = list(results_by_suite_id.get(suite_id, []))
        for cid in children_of.get(suite_id, []):
            out.extend(all_results_for(cid))
        return out

    def render_suite(suite_id: int | None) -> str:
        if suite_id is None:
            name = "—"
            all_r = list(results_by_suite_id.get(None, []))
            child_ids: list[int] = []
            direct_results = all_r
        else:
            suite = suites_by_id[suite_id]
            name = suite.name
            all_r = all_results_for(suite_id)
            child_ids = children_of.get(suite_id, [])
            direct_results = results_by_suite_id.get(suite_id, [])
        sc = {"passed": 0, "failed": 0, "blocked": 0, "not_run": 0}
        suite_dur = 0
        suite_at = None
        for r in all_r:
            sc[r.status] = sc.get(r.status, 0) + 1
            if r.duration_ms:
                suite_dur += r.duration_ms
            if r.executed_at and (suite_at is None or r.executed_at > suite_at):
                suite_at = r.executed_at
        suite_ok = sc["failed"] == 0 and sc["blocked"] == 0
        pill_cls = "kp-passed" if suite_ok else "kp-failed"
        sc_pills = (
            f'<span class="kpill kp-passed">✓ {sc["passed"]}</span>'
            f'<span class="kpill kp-failed">✗ {sc["failed"]}</span>'
            f'<span class="kpill kp-blocked">⚠ {sc["blocked"]}</span>'
            f'<span class="kpill kp-not_run">— {sc["not_run"]}</span>'
        )
        dur_str = fmt_dur(suite_dur) if suite_dur else ""
        at_str = fmt_date(suite_at) if suite_at else ""
        meta_parts = [p for p in [dur_str, at_str] if p and p != "—"]
        suite_meta = " · ".join(meta_parts)
        # Children suites first, then direct test cases
        children_html = "\n    ".join(render_suite(cid) for cid in child_ids)
        results_html = "\n    ".join(render_result(r) for r in direct_results)
        body_parts = [p for p in [children_html, results_html] if p]
        body_inner = "\n    ".join(body_parts)
        err_cls = " suite-err" if not suite_ok else ""
        return f"""<details class="suite{err_cls}" open>
    <summary class="suite-hd">
      <span class="suite-chev">▶</span>
      <span class="kpill {pill_cls}">Suite</span>
      <span class="suite-name">{e(name)}</span>
      <div class="suite-counts">{sc_pills}</div>
      {f'<span class="suite-meta">{e(suite_meta)}</span>' if suite_meta else ""}
    </summary>
    <div class="suite-body">
    {body_inner}
    </div>
  </details>"""

    suites_html = "\n  ".join(render_suite(rid) for rid in root_order)
    tz_label = user_zone_label(current)
    generated_at = datetime.now(timezone.utc).astimezone(user_tz).strftime("%-d %b %Y, %H:%M:%S")
    generated = f"{generated_at} {tz_label} by {e(current.username)}"
    overall_passed = total > 0 and counts["failed"] == 0 and counts["blocked"] == 0
    header_class = "pass" if overall_passed else "fail"

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<meta name="color-scheme" content="light">
<title>{e(ex.title)} — {e(app_settings.app_name)}</title>
<style>
:root{{color-scheme:light}}
*{{box-sizing:border-box;margin:0;padding:0}}
html,body{{color-scheme:light}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f9fafb;color:#111827;font-size:14px;line-height:1.5}}
header{{padding:22px 40px 20px;border-bottom:1px solid}}
header.fail{{background:#fff8f9;border-bottom-color:#fecdd3}}
header.pass{{background:#f0fdf4;border-bottom-color:#bbf7d0}}
.brand{{display:inline-flex;align-items:center;gap:8px;font-size:20px;font-weight:700;letter-spacing:-.01em;color:#e11d48;margin-bottom:12px}}
.brand svg{{display:block}}
header h1{{font-size:22px;font-weight:700;color:#111827;margin-bottom:8px;line-height:1.25}}
.hmeta{{font-size:12px;color:#6b7280;display:flex;flex-wrap:wrap;gap:0 6px;margin-bottom:14px}}
.hmeta span+span::before{{content:"·";margin-right:6px;color:#cbd5e1}}
.hbottom{{display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:10px}}
.hstats{{display:flex;gap:7px;flex-wrap:wrap}}
.hs{{font-size:13px;font-weight:700;padding:5px 14px;border-radius:6px}}
.hs.passed{{background:#d1fae5;color:#065f46}}
.hs.failed{{background:#fee2e2;color:#991b1b}}
.hs.blocked{{background:#fef3c7;color:#92400e}}
.hs.notrun{{background:#f3f4f6;color:#6b7280}}
.hextras{{display:flex;gap:14px;align-items:center;font-size:12px;flex-wrap:wrap}}
.hlink{{color:#e11d48;text-decoration:none;font-weight:600}}
.hlink:hover{{text-decoration:underline}}
.hatts{{color:#9ca3af;display:flex;gap:6px;flex-wrap:wrap;align-items:center}}
.hatts a{{color:#4b5563;text-decoration:none;border:1px solid #e5e7eb;border-radius:4px;padding:2px 8px;background:#fff;font-size:12px}}
.hatts a:hover{{background:#f9fafb;color:#111827;border-color:#d1d5db}}
.hatts{{color:#9ca3af}}
main{{max-width:920px;margin:20px auto;padding:0 24px 48px}}
.toolbar{{display:flex;justify-content:flex-end;gap:6px;margin-bottom:14px}}
.toolbar button{{padding:4px 12px;border:1px solid #e5e7eb;background:#fff;border-radius:6px;font-size:12px;cursor:pointer;color:#374151;transition:background .1s}}
.toolbar button:hover{{background:#f3f4f6;border-color:#d1d5db}}
.kpill{{font-size:11px;padding:2px 7px;border-radius:3px;font-weight:700;flex-shrink:0;white-space:nowrap}}
.kp-passed{{background:#d1fae5;color:#065f46}}
.kp-failed{{background:#fee2e2;color:#991b1b}}
.kp-blocked{{background:#fef3c7;color:#92400e}}
.kp-not_run{{background:#f3f4f6;color:#6b7280}}
details.suite{{background:#fff;border:1px solid #e5e7eb;border-radius:10px;margin-bottom:16px;overflow:hidden}}
details.suite.suite-err{{border-color:#fca5a5}}
summary.suite-hd{{display:flex;align-items:center;gap:10px;padding:11px 16px;cursor:pointer;list-style:none;background:#f9fafb}}
summary.suite-hd::-webkit-details-marker{{display:none}}
details.suite.suite-err>summary.suite-hd{{background:#fff5f5}}
details.suite[open]>summary.suite-hd{{border-bottom:1px solid #e5e7eb}}
.suite-chev{{color:#9ca3af;font-size:10px;transition:transform .15s;flex-shrink:0;pointer-events:none}}
details.suite[open]>summary .suite-chev{{transform:rotate(90deg)}}
.suite-name{{flex:1;font-size:14px;font-weight:700;color:#1f2937}}
.suite-counts{{display:flex;gap:5px;flex-shrink:0}}
.suite-meta{{font-size:11px;color:#9ca3af;white-space:nowrap}}
.suite-body{{padding:10px 12px;display:flex;flex-direction:column;gap:5px}}
details.test{{background:#fff;border:1px solid #e5e7eb;border-radius:8px;overflow:hidden;transition:box-shadow .15s}}
details.test[open]{{box-shadow:0 1px 6px rgba(0,0,0,.07)}}
details.test[data-error]{{border-color:#fca5a5}}
details.test[data-error]>summary.thd{{background:#fff8f8}}
summary.thd{{display:flex;align-items:center;gap:9px;padding:9px 14px;cursor:pointer;list-style:none}}
summary.thd::-webkit-details-marker{{display:none}}
summary.thd:hover{{background:#f9fafb}}
.chevron{{color:#d1d5db;font-size:10px;transition:transform .15s;flex-shrink:0;pointer-events:none}}
details[open]>summary .chevron{{transform:rotate(90deg)}}
.tname{{flex:1;font-weight:500;font-size:13px}}
.tmeta{{font-size:12px;color:#9ca3af;white-space:nowrap}}
.tbody{{border-top:1px solid #f3f4f6;background:#fafafa;padding:10px 14px;display:flex;flex-direction:column;gap:8px}}
.comment{{background:#fef3c7;border-left:3px solid #f59e0b;padding:6px 10px;font-size:12px;border-radius:0 4px 4px 0;color:#78350f}}
.steps{{display:flex;flex-direction:column;gap:4px}}
details.step{{border:1px solid #e5e7eb;border-radius:6px;overflow:hidden;background:#fff}}
details.step.failed{{border-color:#fca5a5}}
details.step.blocked{{border-color:#fcd34d}}
summary.step-hd{{display:flex;align-items:center;gap:7px;padding:6px 10px;cursor:pointer;list-style:none}}
summary.step-hd::-webkit-details-marker{{display:none}}
details.step.failed>summary.step-hd{{background:#fff5f5}}
details.step.blocked>summary.step-hd{{background:#fffbeb}}
.schev{{color:#d1d5db;font-size:9px;transition:transform .15s;flex-shrink:0;pointer-events:none}}
details.step[open] .schev{{transform:rotate(90deg)}}
.snum{{font-size:11px;color:#9ca3af;min-width:18px;flex-shrink:0}}
.saction{{flex:1;font-size:13px;color:#1f2937}}
.sbody{{border-top:1px solid #f3f4f6}}
.sexpected{{padding:6px 12px 6px 34px;font-size:12px;color:#4b5563;background:#f9fafb;display:flex;gap:6px}}
.slabel{{font-size:10px;font-weight:700;color:#9ca3af;text-transform:uppercase;letter-spacing:.04em;flex-shrink:0;padding-top:2px}}
.scomment{{padding:6px 12px 6px 34px;font-size:12px;color:#92400e;background:#fffbeb;border-top:1px solid #fef3c7}}
.slog{{margin:0;padding:10px 14px;background:#f8fafc;color:#334155;font-family:'Monaco','Consolas','Liberation Mono',monospace;font-size:11px;white-space:pre-wrap;overflow-x:auto;border-top:1px solid #e2e8f0;line-height:1.65}}
footer{{text-align:center;padding:20px;font-size:11px;color:#9ca3af;border-top:1px solid #e5e7eb;margin-top:16px}}
</style>
</head>
<body>
<header class="{header_class}">
  <span class="brand">
    <svg width="26" height="26" viewBox="0 0 32 32" aria-hidden="true">
      <g>
        <rect x="6"  y="8"   width="18" height="18" rx="3.2" fill="#fda4af" transform="rotate(-14 15 17)"/>
        <rect x="7"  y="7.5" width="18" height="18" rx="3.2" fill="#f43f5e" transform="rotate(-2 16 16.5)"/>
        <rect x="8"  y="7"   width="18" height="18" rx="3.2" fill="#e11d48" transform="rotate(11 17 16)"/>
      </g>
      <path d="M13.7 16.6l3 3 6.6-7.4" stroke="#fff" stroke-width="2.6"
        fill="none" stroke-linecap="round" stroke-linejoin="round" transform="rotate(11 17 16)" />
    </svg>
    {e(app_settings.app_name)}
  </span>
  <h1>{e(ex.title)}</h1>
  <div class="hmeta">{meta_html}</div>
  <div class="hbottom">
    <div class="hstats">
      <span class="hs passed">✓ {counts["passed"]} passed</span>
      <span class="hs failed">✗ {counts["failed"]} failed</span>
      <span class="hs blocked">⚠ {counts["blocked"]} blocked</span>
      <span class="hs notrun">— {counts["not_run"]} not run</span>
    </div>
    {hextras_html}
  </div>
</header>
<main>
  <div class="toolbar">
    <button onclick="document.querySelectorAll('details').forEach(d=>d.open=true)">Expand all</button>
    <button onclick="document.querySelectorAll('details').forEach(d=>{{if(!d.dataset.error)d.open=false}})">Collapse passing</button>
    <button onclick="document.querySelectorAll('details').forEach(d=>d.open=false)">Collapse all</button>
  </div>
  {suites_html}
</main>
<footer>Generated {generated} · {e(app_settings.app_name)} · {total} test{"s" if total!=1 else ""}</footer>
</body>
</html>"""

    filename = f"execution_{id}_{ex.title.replace(' ', '_')}.html"
    return StreamingResponse(
        io.BytesIO(html.encode("utf-8")),
        media_type="text/html; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@projects_router.get("/{id}/cases/export/xlsx")
def export_cases_xlsx(
    id: int,
    db: Session = Depends(get_db),
    current: User = Depends(require_project_access),
):
    from testjam.models.project import Project
    project = db.get(Project, id)
    suites = db.query(TestSuite).filter(TestSuite.project_id == id, TestSuite.parent_suite_id == None).all()  # noqa: E711

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    # ── Palette ──────────────────────────────────────────────────────────────
    NAVY   = "1E3A5F"
    WHITE  = "FFFFFF"
    GRAY50 = "F9FAFB"
    GRAY200= "E5E7EB"

    thin_border = Border(
        left=Side(style="thin", color=GRAY200),
        right=Side(style="thin", color=GRAY200),
        top=Side(style="thin", color=GRAY200),
        bottom=Side(style="thin", color=GRAY200),
    )

    # ── Title row ────────────────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"Test Cases — {project.name if project else ''}"
    title_cell.font = Font(bold=True, size=13, color=WHITE)
    title_cell.fill = PatternFill("solid", fgColor=NAVY)
    title_cell.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:H2")
    generated_at = format_in_user_zone(
        datetime.now(timezone.utc), current, "%Y-%m-%d %H:%M %Z",
    )
    ws["A2"].value = f"Exported {generated_at} by {current.username}"
    ws["A2"].font = Font(size=8, color="6B7280")
    ws["A2"].alignment = Alignment(indent=1)
    ws.row_dimensions[2].height = 14

    # ── Header row ───────────────────────────────────────────────────────────
    HEADERS = ["Suite", "Test Case", "Preconditions", "Description", "Step #", "Type", "Action", "Expected Result"]
    for col, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, size=9, color=WHITE)
        cell.fill = PatternFill("solid", fgColor="374151")  # gray-700
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        cell.border = thin_border
    ws.row_dimensions[3].height = 16

    STEP_TYPE_COLORS = {"setup": "DBEAFE", "teardown": "FFEDD5", "action": "F9FAFB"}

    data_row = 4
    alt = False

    for suite in suites:
        cases = db.query(TestCase).filter(TestCase.suite_id == suite.id).all()
        for tc in cases:
            steps = sorted(tc.steps, key=lambda s: s.order) if tc.steps else []
            row_count = max(len(steps), 1)
            row_bg = GRAY50 if alt else WHITE
            alt = not alt

            for i in range(row_count):
                step = steps[i] if i < len(steps) else None
                values = [
                    suite.name if i == 0 else "",
                    tc.name if i == 0 else "",
                    tc.preconditions or "" if i == 0 else "",
                    tc.description or "" if i == 0 else "",
                    step.order if step else "",
                    step.step_type if step else "",
                    step.action if step else "",
                    step.expected_result or "" if step else "",
                ]
                step_bg = STEP_TYPE_COLORS.get(step.step_type, row_bg) if step else row_bg

                for col, val in enumerate(values, start=1):
                    cell = ws.cell(row=data_row, column=col, value=val)
                    bg = step_bg if col >= 5 else row_bg
                    cell.fill = PatternFill("solid", fgColor=bg)
                    cell.font = Font(size=9, bold=(col <= 2 and i == 0))
                    cell.alignment = Alignment(
                        vertical="top",
                        wrap_text=(col in (3, 4, 7, 8)),
                        horizontal="center" if col == 5 else "left",
                    )
                    cell.border = thin_border
                ws.row_dimensions[data_row].height = 15
                data_row += 1

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = [20, 28, 22, 30, 6, 10, 40, 30]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"cases_{project.name.replace(' ', '_') if project else id}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
