import html as html_lib
import io
import os
import shutil
import xml.etree.ElementTree as ET
from collections import defaultdict
from datetime import datetime, timezone
from unicodedata import normalize
from urllib.parse import urlparse

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from testjam.auth.dependencies import AuthContext, get_current_user, require_project_access, require_project_access_ctx
from testjam.database import get_db
from testjam.models.execution import ExecutionAttachment, ResultAttachment, TestExecution, TestResult, TestStepResult
from testjam.models.testcase import TestCase, TestSuite
from testjam.models.user import User
from testjam.schemas.execution import (
    BulkResultCreate, BulkResultResponse,
    ExecutionAttachmentOut,
    TestExecutionCreate, TestExecutionOut, TestExecutionUpdate,
    TestResultCreate, TestResultOut, TestResultUpdate,
    TestStepResultUpdate, TestStepResultOut,
    ExecutionSummary,
)
from testjam.schemas.testcase import AttachmentOut

UPLOAD_DIR = "/app/uploads/results"
EXECUTION_UPLOAD_DIR = "/app/uploads/executions"

projects_router = APIRouter(prefix="/projects", tags=["TestExecutions"])
executions_router = APIRouter(prefix="/executions", tags=["TestExecutions"])
results_router = APIRouter(prefix="/results", tags=["TestResults"])


def _compute_summary(execution: TestExecution) -> ExecutionSummary:
    counts = {"passed": 0, "failed": 0, "blocked": 0, "not_run": 0}
    for r in execution.results:
        counts[r.status] = counts.get(r.status, 0) + 1
    return ExecutionSummary(total=len(execution.results), **counts)


def _execution_out(ex: TestExecution) -> TestExecutionOut:
    data = TestExecutionOut.model_validate(ex)
    data.summary = _compute_summary(ex)
    data.attachments = [ExecutionAttachmentOut.model_validate(a) for a in ex.attachments]
    return data


# ─── Executions ───────────────────────────────────────────────────────────────

@projects_router.get("/{id}/executions", response_model=list[TestExecutionOut])
def list_executions(
    id: int,
    type: str | None = None,
    status: str | None = None,
    assigned_to_id: int | None = None,
    skip: int = 0,
    limit: int = 50,
    db: Session = Depends(get_db),
    _: User = Depends(require_project_access),
):
    limit = min(limit, 200)
    q = db.query(TestExecution).filter(TestExecution.project_id == id)
    if type:
        q = q.filter(TestExecution.type == type)
    if status:
        q = q.filter(TestExecution.status == status)
    if assigned_to_id is not None:
        q = q.filter(TestExecution.assigned_to_id == assigned_to_id)
    rows = q.order_by(TestExecution.created_at.desc()).offset(skip).limit(limit).all()
    return [_execution_out(ex) for ex in rows]


@projects_router.post("/{id}/executions", response_model=TestExecutionOut, status_code=status.HTTP_201_CREATED)
def create_execution(
    id: int,
    body: TestExecutionCreate,
    db: Session = Depends(get_db),
    ctx: AuthContext = Depends(require_project_access_ctx),
):
    data = body.model_dump(exclude={"test_case_ids"})
    data["project_id"] = id
    data["created_by_id"] = ctx.user.id
    data["token_name"] = ctx.token_name
    if body.type == "manual" and not data.get("triggered_by"):
        data["triggered_by"] = ctx.user.username
    ex = TestExecution(**data, started_at=datetime.now(timezone.utc))
    db.add(ex)
    db.flush()
    for tc_id in body.test_case_ids:
        db.add(TestResult(execution_id=ex.id, test_case_id=tc_id, status="not_run"))
    db.commit()
    db.refresh(ex)
    return _execution_out(ex)


@executions_router.get("/{id}", response_model=TestExecutionOut)
def get_execution(id: int, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    ex = db.get(TestExecution, id)
    if not ex:
        raise HTTPException(status_code=404, detail="Not found")
    return _execution_out(ex)


@executions_router.put("/{id}", response_model=TestExecutionOut)
def update_execution(id: int, body: TestExecutionUpdate, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    ex = db.get(TestExecution, id)
    if not ex:
        raise HTTPException(status_code=404, detail="Not found")
    for field, value in body.model_dump(exclude_none=True).items():
        setattr(ex, field, value)
    db.commit()
    db.refresh(ex)
    return _execution_out(ex)


@executions_router.delete("/{id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_execution(id: int, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    ex = db.get(TestExecution, id)
    if not ex:
        raise HTTPException(status_code=404, detail="Not found")
    db.delete(ex)
    db.commit()


@executions_router.get("/{id}/export/xlsx")
def export_execution_xlsx(id: int, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    ex = db.get(TestExecution, id)
    if not ex:
        raise HTTPException(status_code=404, detail="Not found")

    wb = openpyxl.Workbook()

    # Summary sheet
    ws_sum = wb.active
    ws_sum.title = "Summary"
    launched_by = ex.token_name or (ex.created_by.username if ex.created_by else ex.triggered_by or "")
    ws_sum.append(["Title", ex.title])
    ws_sum.append(["Status", ex.status])
    ws_sum.append(["Type", ex.type])
    ws_sum.append(["Environment", ex.environment or ""])
    ws_sum.append(["Version", ex.version or ""])
    ws_sum.append(["Launched by", launched_by])
    ws_sum.append(["Started at", ex.started_at.isoformat() if ex.started_at else ""])
    ws_sum.append(["Finished at", ex.finished_at.isoformat() if ex.finished_at else ""])

    # Results sheet
    ws_res = wb.create_sheet("Results")
    ws_res.append(["Test Case", "Status", "Executed by", "Executed at", "Duration (ms)", "Comment"])
    for r in ex.results:
        tc_name = r.test_case.name if r.test_case else str(r.test_case_id)
        ws_res.append([
            tc_name,
            r.status,
            r.executed_by or "",
            r.executed_at.isoformat() if r.executed_at else "",
            r.duration_ms or "",
            r.comment or "",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"execution_{id}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@executions_router.get("/{id}/export/html")
def export_execution_html(id: int, request: Request, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    ex = db.get(TestExecution, id)
    if not ex:
        raise HTTPException(status_code=404, detail="Not found")

    def e(s): return html_lib.escape(str(s)) if s is not None else ""

    def fmt_date(dt):
        if not dt:
            return "—"
        ms = dt.microsecond // 1000
        return dt.strftime("%-d %b %Y, %H:%M:%S") + f".{ms:03d}"

    def fmt_dur(ms):
        if ms is None: return ""
        return f"{ms}ms" if ms < 1000 else f"{ms/1000:.1f}s"

    # Derive frontend execution URL from Referer header
    exec_url = ""
    referer = request.headers.get("referer", "")
    if referer:
        try:
            p = urlparse(referer)
            if p.scheme and p.netloc:
                exec_url = f"{p.scheme}://{p.netloc}/executions/{id}/run"
        except Exception:
            pass

    launched_by = ex.token_name or (ex.created_by.username if ex.created_by else ex.triggered_by or "—")
    counts = {"passed": 0, "failed": 0, "blocked": 0, "not_run": 0}
    for r in ex.results:
        counts[r.status] = counts.get(r.status, 0) + 1
    total = sum(counts.values())

    # Build compact header meta line
    meta_items = []
    project_name = ex.project.name if ex.project else ""
    if project_name:
        meta_items.append(e(project_name))
    if ex.environment:
        meta_items.append(e(ex.environment))
    if ex.version:
        meta_items.append(f"v{e(ex.version)}")
    if launched_by and launched_by != "—":
        meta_items.append(f"by {e(launched_by)}")
    dates_str = fmt_date(ex.started_at)
    if ex.finished_at:
        dates_str += f" → {fmt_date(ex.finished_at)}"
    if dates_str and dates_str != "—":
        meta_items.append(dates_str)
    meta_html = "".join(f'<span>{m}</span>' for m in meta_items)

    # Header extras: execution URL + attachments
    hextras = []
    if exec_url:
        hextras.append(f'<a class="hlink" href="{e(exec_url)}" target="_blank">↗ View in Testjam</a>')
    atts = ex.attachments or []
    if atts:
        att_names = " · ".join(e(a.filename) for a in atts)
        hextras.append(f'<span class="hatts">Attachments: {att_names}</span>')
    hextras_html = f'<div class="hextras">{" ".join(hextras)}</div>' if hextras else ""

    # Group results by suite_id; collect all ancestor suites via lazy load
    results_by_suite_id: dict[int | None, list] = defaultdict(list)
    suites_by_id: dict[int, TestSuite] = {}

    for r in ex.results:
        suite = r.test_case.suite if r.test_case else None
        suite_id = suite.id if suite else None
        results_by_suite_id[suite_id].append(r)
        s = suite
        while s is not None:
            if s.id not in suites_by_id:
                suites_by_id[s.id] = s
            s = s.parent

    # parent_id → [child suite ids]
    children_of: dict[int | None, list[int]] = defaultdict(list)
    for sid, suite in suites_by_id.items():
        parent_id = suite.parent_suite_id if suite.parent_suite_id in suites_by_id else None
        children_of[parent_id].append(sid)

    def _root_id(suite_id: int | None) -> int | None:
        if suite_id is None or suite_id not in suites_by_id:
            return None
        pid = suites_by_id[suite_id].parent_suite_id
        return suite_id if (pid is None or pid not in suites_by_id) else _root_id(pid)

    # Preserve root order from ex.results insertion order
    root_order: list[int | None] = []
    seen_roots: set[int | None] = set()
    for r in ex.results:
        suite = r.test_case.suite if r.test_case else None
        rid = _root_id(suite.id if suite else None)
        if rid not in seen_roots:
            seen_roots.add(rid)
            root_order.append(rid)

    STATUS_LABEL = {"passed": "Passed", "failed": "Failed", "blocked": "Blocked", "not_run": "Not run"}

    def render_steps(r) -> str:
        steps = sorted(r.test_case.steps, key=lambda s: s.order) if r.test_case and r.test_case.steps else []
        sr_by_step = {sr.step_id: sr for sr in r.step_results}
        if not steps:
            return ""
        parts = []
        for step in steps:
            sr = sr_by_step.get(step.id)
            st = sr.status if sr else "not_run"
            is_step_error = st in ("failed", "blocked")
            log = sr.log_output if sr else None
            comment = sr.comment if sr else None
            dur = fmt_dur(sr.duration_ms) if sr else ""
            has_body = step.expected_result or comment or log
            open_attr = 'open data-error="1"' if is_step_error else ""
            expected_html = f'<div class="sexpected"><span class="slabel">Expected</span>{e(step.expected_result)}</div>' if step.expected_result else ""
            comment_html = f'<div class="scomment">💬 {e(comment)}</div>' if comment else ""
            log_html = f'<pre class="slog">{e(log)}</pre>' if log else ""
            body_html = f'<div class="sbody">{expected_html}{comment_html}{log_html}</div>' if has_body else ""
            parts.append(f"""<details class="step {e(st)}" {open_attr}>
        <summary class="step-hd">
          <span class="schev">▶</span>
          <span class="snum">{step.order}.</span>
          <span class="stype {e(step.step_type)}">{e(step.step_type)}</span>
          <span class="saction">{e(step.action)}</span>
          <span class="sbadge {e(st)}">{e(STATUS_LABEL.get(st, st))}</span>
          {f'<span class="sdur">{e(dur)}</span>' if dur else ""}
        </summary>
        {body_html}
      </details>""")
        return "\n      ".join(parts)

    def render_result(r) -> str:
        st = r.status
        is_error = st in ("failed", "blocked")
        dur = fmt_dur(r.duration_ms)
        meta_parts = [p for p in [e(r.executed_by), dur, fmt_date(r.executed_at)] if p and p != "—"]
        meta = " · ".join(meta_parts) if meta_parts else ""
        steps_html = render_steps(r)
        comment_html = f'<div class="comment">💬 {e(r.comment)}</div>' if r.comment else ""
        steps_div = f'<div class="steps">{steps_html}</div>' if steps_html else ""
        body = f'<div class="tbody">{comment_html}{steps_div}</div>' if (comment_html or steps_html) else ""
        open_attr = 'open data-error="1"' if is_error else ""
        return f"""<details class="test" {open_attr}>
      <summary class="thd">
        <span class="chevron">▶</span>
        <span class="kpill kp-{e(st)}">Test</span>
        <span class="tname">{e(r.test_case.name if r.test_case else str(r.test_case_id))}</span>
        {f'<span class="tmeta">{e(meta)}</span>' if meta else ""}
      </summary>
      {body}
    </details>"""

    def all_results_for(suite_id: int) -> list:
        out = list(results_by_suite_id.get(suite_id, []))
        for cid in children_of.get(suite_id, []):
            out.extend(all_results_for(cid))
        return out

    def render_suite(suite_id: int | None) -> str:
        if suite_id is None:
            name = "—"
            all_r = list(results_by_suite_id.get(None, []))
            child_ids: list[int] = []
            direct_results = all_r
        else:
            suite = suites_by_id[suite_id]
            name = suite.name
            all_r = all_results_for(suite_id)
            child_ids = children_of.get(suite_id, [])
            direct_results = results_by_suite_id.get(suite_id, [])
        sc = {"passed": 0, "failed": 0, "blocked": 0, "not_run": 0}
        suite_dur = 0
        suite_at = None
        for r in all_r:
            sc[r.status] = sc.get(r.status, 0) + 1
            if r.duration_ms:
                suite_dur += r.duration_ms
            if r.executed_at and (suite_at is None or r.executed_at > suite_at):
                suite_at = r.executed_at
        suite_ok = sc["failed"] == 0 and sc["blocked"] == 0
        pill_cls = "kp-passed" if suite_ok else "kp-failed"
        sc_pills = (
            f'<span class="kpill kp-passed">✓ {sc["passed"]}</span>'
            f'<span class="kpill kp-failed">✗ {sc["failed"]}</span>'
            f'<span class="kpill kp-blocked">⚠ {sc["blocked"]}</span>'
            f'<span class="kpill kp-not_run">— {sc["not_run"]}</span>'
        )
        dur_str = fmt_dur(suite_dur) if suite_dur else ""
        at_str = fmt_date(suite_at) if suite_at else ""
        meta_parts = [p for p in [dur_str, at_str] if p and p != "—"]
        suite_meta = " · ".join(meta_parts)
        # Children suites first, then direct test cases
        children_html = "\n    ".join(render_suite(cid) for cid in child_ids)
        results_html = "\n    ".join(render_result(r) for r in direct_results)
        body_parts = [p for p in [children_html, results_html] if p]
        body_inner = "\n    ".join(body_parts)
        err_cls = " suite-err" if not suite_ok else ""
        return f"""<details class="suite{err_cls}" open>
    <summary class="suite-hd">
      <span class="suite-chev">▶</span>
      <span class="kpill {pill_cls}">Suite</span>
      <span class="suite-name">{e(name)}</span>
      <div class="suite-counts">{sc_pills}</div>
      {f'<span class="suite-meta">{e(suite_meta)}</span>' if suite_meta else ""}
    </summary>
    <div class="suite-body">
    {body_inner}
    </div>
  </details>"""

    suites_html = "\n  ".join(render_suite(rid) for rid in root_order)
    generated = datetime.now().strftime("%-d %b %Y, %H:%M:%S")

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{e(ex.title)} — Testjam</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f9fafb;color:#111827;font-size:14px;line-height:1.5}}
header{{background:#fff8f9;border-bottom:1px solid #fecdd3;padding:22px 40px 20px}}
.brand{{display:inline-flex;align-items:center;gap:8px;font-size:20px;font-weight:700;letter-spacing:-.01em;color:#e11d48;margin-bottom:12px}}
.brand svg{{display:block}}
header h1{{font-size:22px;font-weight:700;color:#111827;margin-bottom:8px;line-height:1.25}}
.hmeta{{font-size:12px;color:#6b7280;display:flex;flex-wrap:wrap;gap:0 6px;margin-bottom:14px}}
.hmeta span+span::before{{content:"·";margin-right:6px;color:#fca5a5}}
.hbottom{{display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:10px}}
.hstats{{display:flex;gap:7px;flex-wrap:wrap}}
.hs{{font-size:13px;font-weight:700;padding:5px 14px;border-radius:6px}}
.hs.passed{{background:#d1fae5;color:#065f46}}
.hs.failed{{background:#fee2e2;color:#991b1b}}
.hs.blocked{{background:#fef3c7;color:#92400e}}
.hs.notrun{{background:#f3f4f6;color:#6b7280}}
.hextras{{display:flex;gap:14px;align-items:center;font-size:12px}}
.hlink{{color:#e11d48;text-decoration:none;font-weight:600}}
.hlink:hover{{text-decoration:underline}}
.hatts{{color:#9ca3af}}
main{{max-width:920px;margin:20px auto;padding:0 24px 48px}}
.toolbar{{display:flex;justify-content:flex-end;gap:6px;margin-bottom:14px}}
.toolbar button{{padding:4px 12px;border:1px solid #e5e7eb;background:#fff;border-radius:6px;font-size:12px;cursor:pointer;color:#374151;transition:background .1s}}
.toolbar button:hover{{background:#f3f4f6;border-color:#d1d5db}}
.kpill{{font-size:11px;padding:2px 7px;border-radius:3px;font-weight:700;flex-shrink:0;white-space:nowrap}}
.kp-passed{{background:#d1fae5;color:#065f46}}
.kp-failed{{background:#fee2e2;color:#991b1b}}
.kp-blocked{{background:#fef3c7;color:#92400e}}
.kp-not_run{{background:#f3f4f6;color:#6b7280}}
details.suite{{background:#fff;border:1px solid #e5e7eb;border-radius:10px;margin-bottom:16px;overflow:hidden}}
details.suite.suite-err{{border-color:#fca5a5}}
summary.suite-hd{{display:flex;align-items:center;gap:10px;padding:11px 16px;cursor:pointer;list-style:none;background:#f9fafb}}
summary.suite-hd::-webkit-details-marker{{display:none}}
details.suite.suite-err>summary.suite-hd{{background:#fff5f5}}
details.suite[open]>summary.suite-hd{{border-bottom:1px solid #e5e7eb}}
.suite-chev{{color:#9ca3af;font-size:10px;transition:transform .15s;flex-shrink:0;pointer-events:none}}
details.suite[open]>summary .suite-chev{{transform:rotate(90deg)}}
.suite-name{{flex:1;font-size:14px;font-weight:700;color:#1f2937}}
.suite-counts{{display:flex;gap:5px;flex-shrink:0}}
.suite-meta{{font-size:11px;color:#9ca3af;white-space:nowrap}}
.suite-body{{padding:10px 12px;display:flex;flex-direction:column;gap:5px}}
details.test{{background:#fff;border:1px solid #e5e7eb;border-radius:8px;overflow:hidden;transition:box-shadow .15s}}
details.test[open]{{box-shadow:0 1px 6px rgba(0,0,0,.07)}}
details.test[data-error]{{border-color:#fca5a5}}
details.test[data-error]>summary.thd{{background:#fff8f8}}
summary.thd{{display:flex;align-items:center;gap:9px;padding:9px 14px;cursor:pointer;list-style:none}}
summary.thd::-webkit-details-marker{{display:none}}
summary.thd:hover{{background:#f9fafb}}
.chevron{{color:#d1d5db;font-size:10px;transition:transform .15s;flex-shrink:0;pointer-events:none}}
details[open]>summary .chevron{{transform:rotate(90deg)}}
.tname{{flex:1;font-weight:500;font-size:13px}}
.tmeta{{font-size:12px;color:#9ca3af;white-space:nowrap}}
.tbody{{border-top:1px solid #f3f4f6;background:#fafafa;padding:10px 14px;display:flex;flex-direction:column;gap:8px}}
.comment{{background:#fef3c7;border-left:3px solid #f59e0b;padding:6px 10px;font-size:12px;border-radius:0 4px 4px 0;color:#78350f}}
.steps{{display:flex;flex-direction:column;gap:4px}}
details.step{{border:1px solid #e5e7eb;border-radius:6px;overflow:hidden;background:#fff}}
details.step.failed{{border-color:#fca5a5}}
details.step.blocked{{border-color:#fcd34d}}
summary.step-hd{{display:flex;align-items:center;gap:7px;padding:6px 10px;cursor:pointer;list-style:none}}
summary.step-hd::-webkit-details-marker{{display:none}}
details.step.failed>summary.step-hd{{background:#fff5f5}}
details.step.blocked>summary.step-hd{{background:#fffbeb}}
.schev{{color:#d1d5db;font-size:9px;transition:transform .15s;flex-shrink:0;pointer-events:none}}
details.step[open] .schev{{transform:rotate(90deg)}}
.snum{{font-size:11px;color:#9ca3af;min-width:18px;flex-shrink:0}}
.stype{{font-size:10px;padding:1px 5px;border-radius:3px;font-weight:700;flex-shrink:0}}
.stype.setup{{background:#dbeafe;color:#1e40af}}
.stype.teardown{{background:#ffedd5;color:#9a3412}}
.stype.action{{background:#f3f4f6;color:#4b5563}}
.saction{{flex:1;font-size:13px;color:#1f2937}}
.sbadge{{font-size:10px;font-weight:700;padding:1px 5px;border-radius:3px;flex-shrink:0}}
.sbadge.passed{{background:#d1fae5;color:#065f46}}
.sbadge.failed{{background:#fee2e2;color:#991b1b}}
.sbadge.blocked{{background:#fef3c7;color:#92400e}}
.sbadge.not_run{{background:#f3f4f6;color:#9ca3af}}
.sdur{{font-size:11px;color:#9ca3af;min-width:36px;text-align:right;flex-shrink:0}}
.sbody{{border-top:1px solid #f3f4f6}}
.sexpected{{padding:6px 12px 6px 34px;font-size:12px;color:#4b5563;background:#f9fafb;display:flex;gap:6px}}
.slabel{{font-size:10px;font-weight:700;color:#9ca3af;text-transform:uppercase;letter-spacing:.04em;flex-shrink:0;padding-top:2px}}
.scomment{{padding:6px 12px 6px 34px;font-size:12px;color:#92400e;background:#fffbeb;border-top:1px solid #fef3c7}}
.slog{{margin:0;padding:10px 14px;background:#f8fafc;color:#334155;font-family:'Monaco','Consolas','Liberation Mono',monospace;font-size:11px;white-space:pre-wrap;overflow-x:auto;border-top:1px solid #e2e8f0;line-height:1.65}}
footer{{text-align:center;padding:20px;font-size:11px;color:#9ca3af;border-top:1px solid #e5e7eb;margin-top:16px}}
</style>
</head>
<body>
<header>
  <span class="brand">
    <svg width="26" height="26" viewBox="0 0 32 32" aria-hidden="true">
      <g>
        <rect x="6"  y="8"   width="18" height="18" rx="3.2" fill="#fda4af" transform="rotate(-14 15 17)"/>
        <rect x="7"  y="7.5" width="18" height="18" rx="3.2" fill="#f43f5e" transform="rotate(-2 16 16.5)"/>
        <rect x="8"  y="7"   width="18" height="18" rx="3.2" fill="#e11d48" transform="rotate(11 17 16)"/>
      </g>
      <path d="M13.7 16.6l3 3 6.6-7.4" stroke="#fff" stroke-width="2.6"
        fill="none" stroke-linecap="round" stroke-linejoin="round" transform="rotate(11 17 16)" />
    </svg>
    Testjam
  </span>
  <h1>{e(ex.title)}</h1>
  <div class="hmeta">{meta_html}</div>
  <div class="hbottom">
    <div class="hstats">
      <span class="hs passed">✓ {counts["passed"]} passed</span>
      <span class="hs failed">✗ {counts["failed"]} failed</span>
      <span class="hs blocked">⚠ {counts["blocked"]} blocked</span>
      <span class="hs notrun">— {counts["not_run"]} not run</span>
    </div>
    {hextras_html}
  </div>
</header>
<main>
  <div class="toolbar">
    <button onclick="document.querySelectorAll('details').forEach(d=>d.open=true)">Expand all</button>
    <button onclick="document.querySelectorAll('details').forEach(d=>{{if(!d.dataset.error)d.open=false}})">Collapse passing</button>
    <button onclick="document.querySelectorAll('details').forEach(d=>d.open=false)">Collapse all</button>
  </div>
  {suites_html}
</main>
<footer>Generated {generated} · Testjam · {total} test{"s" if total!=1 else ""}</footer>
</body>
</html>"""

    filename = f"execution_{id}_{ex.title.replace(' ', '_')}.html"
    return StreamingResponse(
        io.BytesIO(html.encode("utf-8")),
        media_type="text/html; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@projects_router.get("/{id}/cases/export/xlsx")
def export_cases_xlsx(id: int, db: Session = Depends(get_db), _: User = Depends(require_project_access)):
    from testjam.models.project import Project
    project = db.get(Project, id)
    suites = db.query(TestSuite).filter(TestSuite.project_id == id, TestSuite.parent_suite_id == None).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    # ── Palette ──────────────────────────────────────────────────────────────
    NAVY   = "1E3A5F"
    WHITE  = "FFFFFF"
    GRAY50 = "F9FAFB"
    GRAY200= "E5E7EB"
    BLUE50 = "EFF6FF"
    SUITE_BG = "1E3A5F"

    thin_border = Border(
        left=Side(style="thin", color=GRAY200),
        right=Side(style="thin", color=GRAY200),
        top=Side(style="thin", color=GRAY200),
        bottom=Side(style="thin", color=GRAY200),
    )

    # ── Title row ────────────────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"Test Cases — {project.name if project else ''}"
    title_cell.font = Font(bold=True, size=13, color=WHITE)
    title_cell.fill = PatternFill("solid", fgColor=NAVY)
    title_cell.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:H2")
    ws["A2"].value = f"Exported {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(size=8, color="6B7280")
    ws["A2"].alignment = Alignment(indent=1)
    ws.row_dimensions[2].height = 14

    # ── Header row ───────────────────────────────────────────────────────────
    HEADERS = ["Suite", "Test Case", "Preconditions", "Description", "Step #", "Type", "Action", "Expected Result"]
    for col, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, size=9, color=WHITE)
        cell.fill = PatternFill("solid", fgColor="374151")  # gray-700
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        cell.border = thin_border
    ws.row_dimensions[3].height = 16

    STEP_TYPE_COLORS = {"setup": "DBEAFE", "teardown": "FFEDD5", "action": "F9FAFB"}

    data_row = 4
    alt = False

    for suite in suites:
        cases = db.query(TestCase).filter(TestCase.suite_id == suite.id).all()
        for tc in cases:
            steps = sorted(tc.steps, key=lambda s: s.order) if tc.steps else []
            row_count = max(len(steps), 1)
            row_bg = GRAY50 if alt else WHITE
            alt = not alt

            for i in range(row_count):
                step = steps[i] if i < len(steps) else None
                values = [
                    suite.name if i == 0 else "",
                    tc.name if i == 0 else "",
                    tc.preconditions or "" if i == 0 else "",
                    tc.description or "" if i == 0 else "",
                    step.order if step else "",
                    step.step_type if step else "",
                    step.action if step else "",
                    step.expected_result or "" if step else "",
                ]
                step_bg = STEP_TYPE_COLORS.get(step.step_type, row_bg) if step else row_bg

                for col, val in enumerate(values, start=1):
                    cell = ws.cell(row=data_row, column=col, value=val)
                    bg = step_bg if col >= 5 else row_bg
                    cell.fill = PatternFill("solid", fgColor=bg)
                    cell.font = Font(size=9, bold=(col <= 2 and i == 0))
                    cell.alignment = Alignment(
                        vertical="top",
                        wrap_text=(col in (3, 4, 7, 8)),
                        horizontal="center" if col == 5 else "left",
                    )
                    cell.border = thin_border
                ws.row_dimensions[data_row].height = 15
                data_row += 1

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = [20, 28, 22, 30, 6, 10, 40, 30]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"cases_{project.name.replace(' ', '_') if project else id}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─── Execution attachments ────────────────────────────────────────────────────

@executions_router.get("/{id}/attachments", response_model=list[ExecutionAttachmentOut])
def list_execution_attachments(id: int, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return db.query(ExecutionAttachment).filter(ExecutionAttachment.execution_id == id).all()


@executions_router.post("/{id}/attachments", response_model=ExecutionAttachmentOut, status_code=status.HTTP_201_CREATED)
def upload_execution_attachment(
    id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current: User = Depends(get_current_user),
):
    if not db.get(TestExecution, id):
        raise HTTPException(status_code=404, detail="Not found")
    dest_dir = os.path.join(EXECUTION_UPLOAD_DIR, str(id))
    os.makedirs(dest_dir, exist_ok=True)
    dest_path = os.path.join(dest_dir, file.filename)
    with open(dest_path, "wb") as f:
        shutil.copyfileobj(file.file, f)
    att = ExecutionAttachment(
        execution_id=id,
        filename=file.filename,
        content_type=file.content_type,
        size_bytes=os.path.getsize(dest_path),
        file_path=dest_path,
        uploaded_by=current.username,
    )
    db.add(att)
    db.commit()
    db.refresh(att)
    return att


@executions_router.delete("/{id}/attachments/{attachment_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_execution_attachment(id: int, attachment_id: int, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    att = db.query(ExecutionAttachment).filter(ExecutionAttachment.id == attachment_id, ExecutionAttachment.execution_id == id).first()
    if not att:
        raise HTTPException(status_code=404, detail="Not found")
    if os.path.exists(att.file_path):
        os.remove(att.file_path)
    db.delete(att)
    db.commit()


# ─── Results ──────────────────────────────────────────────────────────────────

@executions_router.get("/{id}/results", response_model=list[TestResultOut])
def list_results(id: int, status: str | None = None, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    q = db.query(TestResult).filter(TestResult.execution_id == id)
    if status:
        q = q.filter(TestResult.status == status)
    results = q.order_by(TestResult.id).all()
    out = []
    for r in results:
        ro = TestResultOut.model_validate(r)
        ro.test_case_title = r.test_case.name if r.test_case else None
        out.append(ro)
    return out


@executions_router.post("/{id}/results", response_model=TestResultOut, status_code=status.HTTP_201_CREATED)
def create_result(id: int, body: TestResultCreate, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    existing = db.query(TestResult).filter_by(execution_id=id, test_case_id=body.test_case_id).first()
    if existing:
        for field, value in body.model_dump(exclude={"step_results"}, exclude_unset=True).items():
            setattr(existing, field, value)
        result = existing
    else:
        result = TestResult(**body.model_dump(exclude={"step_results"}), execution_id=id)
        db.add(result)
    db.flush()
    for sr in body.step_results:
        existing_sr = db.query(TestStepResult).filter_by(test_result_id=result.id, step_id=sr.step_id).first()
        if existing_sr:
            existing_sr.status = sr.status
            existing_sr.comment = sr.comment
            existing_sr.log_output = sr.log_output
        else:
            db.add(TestStepResult(test_result_id=result.id, **sr.model_dump()))
    db.commit()
    db.refresh(result)
    ro = TestResultOut.model_validate(result)
    ro.test_case_title = result.test_case.name if result.test_case else None
    return ro


@executions_router.post("/{id}/results/bulk", response_model=BulkResultResponse)
def bulk_results(id: int, body: BulkResultCreate, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    created = updated = 0
    errors = []
    for item in body.results:
        try:
            existing = db.query(TestResult).filter_by(execution_id=id, test_case_id=item.test_case_id).first()
            if existing:
                for field, value in item.model_dump(exclude={"step_results"}, exclude_none=True).items():
                    setattr(existing, field, value)
                result = existing
                updated += 1
            else:
                result = TestResult(**item.model_dump(exclude={"step_results"}), execution_id=id)
                db.add(result)
                created += 1
            db.flush()
            for sr in item.step_results:
                existing_sr = db.query(TestStepResult).filter_by(test_result_id=result.id, step_id=sr.step_id).first()
                if existing_sr:
                    existing_sr.status = sr.status
                    existing_sr.comment = sr.comment
                else:
                    db.add(TestStepResult(test_result_id=result.id, **sr.model_dump()))
        except Exception as e:
            errors.append({"test_case_id": item.test_case_id, "error": str(e)})
    db.commit()
    return BulkResultResponse(created=created, updated=updated, errors=errors)


@results_router.get("/{id}", response_model=TestResultOut)
def get_result(id: int, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    result = db.get(TestResult, id)
    if not result:
        raise HTTPException(status_code=404, detail="Not found")
    ro = TestResultOut.model_validate(result)
    ro.test_case_title = result.test_case.name if result.test_case else None
    return ro


@results_router.put("/{id}", response_model=TestResultOut)
def update_result(id: int, body: TestResultUpdate, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    result = db.get(TestResult, id)
    if not result:
        raise HTTPException(status_code=404, detail="Not found")
    for field, value in body.model_dump(exclude_none=True).items():
        setattr(result, field, value)
    db.commit()
    db.refresh(result)
    ro = TestResultOut.model_validate(result)
    ro.test_case_title = result.test_case.name if result.test_case else None
    return ro


@results_router.put("/{id}/step-results/{step_result_id}", response_model=TestStepResultOut)
def update_step_result(id: int, step_result_id: int, body: TestStepResultUpdate, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    sr = db.query(TestStepResult).filter(TestStepResult.id == step_result_id, TestStepResult.test_result_id == id).first()
    if not sr:
        raise HTTPException(status_code=404, detail="Not found")
    for field, value in body.model_dump(exclude_none=True).items():
        setattr(sr, field, value)
    db.commit()
    db.refresh(sr)
    return sr


# ─── Result attachments ───────────────────────────────────────────────────────

@results_router.get("/{id}/attachments", response_model=list[AttachmentOut])
def list_result_attachments(id: int, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return db.query(ResultAttachment).filter(ResultAttachment.result_id == id).all()


@results_router.post("/{id}/attachments", status_code=status.HTTP_201_CREATED)
def upload_result_attachment(
    id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current: User = Depends(get_current_user),
):
    if not db.get(TestResult, id):
        raise HTTPException(status_code=404, detail="Not found")
    dest_dir = os.path.join(UPLOAD_DIR, str(id))
    os.makedirs(dest_dir, exist_ok=True)
    dest_path = os.path.join(dest_dir, file.filename)
    with open(dest_path, "wb") as f:
        shutil.copyfileobj(file.file, f)
    att = ResultAttachment(
        result_id=id,
        filename=file.filename,
        content_type=file.content_type,
        size_bytes=os.path.getsize(dest_path),
        file_path=dest_path,
        uploaded_by=current.username,
    )
    db.add(att)
    db.commit()
    db.refresh(att)
    return att


@results_router.delete("/{id}/attachments/{attachment_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_result_attachment(id: int, attachment_id: int, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    att = db.query(ResultAttachment).filter(ResultAttachment.id == attachment_id, ResultAttachment.result_id == id).first()
    if not att:
        raise HTTPException(status_code=404, detail="Not found")
    if os.path.exists(att.file_path):
        os.remove(att.file_path)
    db.delete(att)
    db.commit()


# ─── Import helpers ───────────────────────────────────────────────────────────

def _normalize(s: str) -> str:
    return normalize("NFC", s).strip().lower()


def _build_result_index(execution: TestExecution) -> dict[str, TestResult]:
    """Return a dict keyed by external_id and by normalized title for fast lookup."""
    index: dict[str, TestResult] = {}
    for r in execution.results:
        tc = r.test_case
        if not tc:
            continue
        if tc.external_id:
            index[_normalize(tc.external_id)] = r
        index[_normalize(tc.name)] = r
    return index


def _rf_status(status_str: str) -> str:
    mapping = {"PASS": "passed", "FAIL": "failed", "SKIP": "blocked", "NOT RUN": "not_run"}
    return mapping.get(status_str.upper(), "not_run")


def _junit_status(tc_elem: ET.Element) -> str:
    if tc_elem.find("failure") is not None or tc_elem.find("error") is not None:
        return "failed"
    if tc_elem.find("skipped") is not None:
        return "blocked"
    return "passed"


def _rf_collect_messages(kw_elem: ET.Element) -> str:
    """Collect all log messages from a keyword element into markdown."""
    lines = []
    for msg in kw_elem.iter("msg"):
        level = msg.get("level", "INFO")
        text = (msg.text or "").strip()
        if text:
            lines.append(f"**[{level}]** {text}")
    return "\n\n".join(lines)


def _rf_kw_duration_ms(kw_elem: ET.Element) -> int | None:
    s = kw_elem.find("status")
    if s is None:
        return None
    start = s.get("starttime") or s.get("start")
    end = s.get("endtime") or s.get("end")
    if start and end:
        try:
            fmt = "%Y%m%d %H:%M:%S.%f"
            delta = datetime.strptime(end, fmt) - datetime.strptime(start, fmt)
            return int(delta.total_seconds() * 1000)
        except ValueError:
            return None
    return None


# ─── Import: JUnit XML ────────────────────────────────────────────────────────

@executions_router.post("/{id}/results/import/junit", response_model=BulkResultResponse)
def import_junit(
    id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    ex = db.get(TestExecution, id)
    if not ex:
        raise HTTPException(status_code=404, detail="Execution not found")

    try:
        tree = ET.parse(file.file)
    except ET.ParseError as e:
        raise HTTPException(status_code=400, detail=f"Invalid XML: {e}")

    root = tree.getroot()
    # Support both <testsuites><testsuite>... and bare <testsuite>...
    if root.tag == "testsuites":
        tc_elements = [tc for ts in root.findall("testsuite") for tc in ts.findall("testcase")]
    elif root.tag == "testsuite":
        tc_elements = root.findall("testcase")
    else:
        raise HTTPException(status_code=400, detail="Root element must be <testsuite> or <testsuites>")

    result_index = _build_result_index(ex)
    created = updated = 0
    errors: list[dict] = []
    now = datetime.now(timezone.utc)

    for tc_elem in tc_elements:
        name = tc_elem.get("name", "")
        classname = tc_elem.get("classname", "")
        # Try matching by classname.name, then just name
        candidates = [
            _normalize(f"{classname}.{name}") if classname else None,
            _normalize(f"{classname}::{name}") if classname else None,
            _normalize(name),
        ]

        result: TestResult | None = None
        for key in candidates:
            if key and key in result_index:
                result = result_index[key]
                break

        if result is None:
            errors.append({"name": name, "classname": classname, "error": "No matching test case found"})
            continue

        new_status = _junit_status(tc_elem)
        failure = tc_elem.find("failure") or tc_elem.find("error")
        comment = failure.get("message", "") if failure is not None else None
        duration_raw = tc_elem.get("time")
        duration_ms = int(float(duration_raw) * 1000) if duration_raw else None

        is_new = result.status == "not_run"
        result.status = new_status
        result.executed_at = now
        result.duration_ms = duration_ms
        if comment:
            result.comment = comment
        if is_new:
            created += 1
        else:
            updated += 1

    db.commit()
    return BulkResultResponse(created=created, updated=updated, errors=errors)


# ─── Import: Robot Framework output.xml ──────────────────────────────────────

@executions_router.post("/{id}/results/import/robotframework", response_model=BulkResultResponse)
def import_robotframework(
    id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    ex = db.get(TestExecution, id)
    if not ex:
        raise HTTPException(status_code=404, detail="Execution not found")

    try:
        tree = ET.parse(file.file)
    except ET.ParseError as e:
        raise HTTPException(status_code=400, detail=f"Invalid XML: {e}")

    root = tree.getroot()
    if root.tag not in ("robot", "suite"):
        raise HTTPException(status_code=400, detail="Root element must be <robot> or <suite>")

    result_index = _build_result_index(ex)
    created = updated = 0
    errors: list[dict] = []
    now = datetime.now(timezone.utc)

    def _process_suite(suite_elem: ET.Element, parent_path: str = "") -> None:
        nonlocal created, updated
        suite_name = suite_elem.get("name", "")
        suite_path = f"{parent_path}.{suite_name}" if parent_path else suite_name

        for test_elem in suite_elem.findall("test"):
            test_name = test_elem.get("name", "")
            full_path = f"{suite_path}.{test_name}"

            candidates = [
                _normalize(full_path),
                _normalize(test_name),
            ]
            result: TestResult | None = None
            for key in candidates:
                if key in result_index:
                    result = result_index[key]
                    break

            if result is None:
                errors.append({"name": test_name, "suite": suite_path, "error": "No matching test case found"})
                continue

            status_elem = test_elem.find("status")
            rf_status = status_elem.get("status", "NOT RUN") if status_elem is not None else "NOT RUN"
            new_status = _rf_status(rf_status)

            # Duration from status element
            duration_ms = None
            if status_elem is not None:
                start = status_elem.get("starttime") or status_elem.get("start")
                end = status_elem.get("endtime") or status_elem.get("end")
                if start and end:
                    try:
                        fmt = "%Y%m%d %H:%M:%S.%f"
                        delta = datetime.strptime(end, fmt) - datetime.strptime(start, fmt)
                        duration_ms = int(delta.total_seconds() * 1000)
                    except ValueError:
                        pass

            is_new = result.status == "not_run"
            result.status = new_status
            result.executed_at = now
            if duration_ms is not None:
                result.duration_ms = duration_ms

            # Collect status message as comment
            if status_elem is not None and status_elem.text:
                result.comment = status_elem.text.strip()

            db.flush()

            # Map keywords to step results if the test case has steps
            tc = result.test_case
            if tc and tc.steps:
                steps = sorted(tc.steps, key=lambda s: (
                    {"setup": 0, "action": 1, "teardown": 2}.get(s.step_type, 1), s.order
                ))
                kw_elems = list(test_elem.findall("kw"))

                for i, (step, kw) in enumerate(zip(steps, kw_elems)):
                    kw_status_elem = kw.find("status")
                    kw_rf_status = kw_status_elem.get("status", "NOT RUN") if kw_status_elem is not None else "NOT RUN"
                    kw_new_status = _rf_status(kw_rf_status)
                    log_output = _rf_collect_messages(kw)
                    kw_duration = _rf_kw_duration_ms(kw)

                    existing_sr = db.query(TestStepResult).filter_by(
                        test_result_id=result.id, step_id=step.id
                    ).first()
                    if existing_sr:
                        existing_sr.status = kw_new_status
                        existing_sr.log_output = log_output or None
                        existing_sr.duration_ms = kw_duration
                    else:
                        db.add(TestStepResult(
                            test_result_id=result.id,
                            step_id=step.id,
                            status=kw_new_status,
                            log_output=log_output or None,
                            duration_ms=kw_duration,
                        ))

            if is_new:
                created += 1
            else:
                updated += 1

        for child_suite in suite_elem.findall("suite"):
            _process_suite(child_suite, suite_path)

    top = root if root.tag == "suite" else root.find("suite")
    if top is not None:
        _process_suite(top)

    db.commit()
    return BulkResultResponse(created=created, updated=updated, errors=errors)
