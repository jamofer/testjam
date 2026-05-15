"""Exports render dates in the requester's timezone and tag the footer."""
import re

import openpyxl


def test_html_export_footer_includes_user_timezone_and_username(auth_client, project_id, case_ids):
    auth_client.put("/api/v1/users/me", json={"timezone": "America/New_York"})
    exec_id = auth_client.post(f"/api/v1/projects/{project_id}/executions", json={
        "title": "TZ-aware run", "type": "manual", "test_case_ids": case_ids,
    }).json()["id"]
    auth_client.post(f"/api/v1/executions/{exec_id}/results", json={
        "test_case_id": case_ids[0], "status": "passed",
    })

    resp = auth_client.get(f"/api/v1/executions/{exec_id}/export/html")

    assert resp.status_code == 200
    footer_match = re.search(r"Generated ([^<·]+)·", resp.text)
    assert footer_match, "Footer Generated stamp missing"
    footer = footer_match.group(1)
    assert "by u" in footer
    assert "EST" in footer or "EDT" in footer


def test_html_export_falls_back_to_utc_when_user_has_no_timezone(auth_client, project_id, case_ids):
    exec_id = auth_client.post(f"/api/v1/projects/{project_id}/executions", json={
        "title": "Default tz", "type": "manual", "test_case_ids": case_ids,
    }).json()["id"]

    resp = auth_client.get(f"/api/v1/executions/{exec_id}/export/html")

    assert resp.status_code == 200
    assert "UTC by u" in resp.text


def test_xlsx_summary_sheet_carries_timezone_label(auth_client, project_id, case_ids, tmp_path):
    auth_client.put("/api/v1/users/me", json={"timezone": "Europe/Madrid"})
    exec_id = auth_client.post(f"/api/v1/projects/{project_id}/executions", json={
        "title": "Madrid run", "type": "manual", "test_case_ids": case_ids,
    }).json()["id"]

    resp = auth_client.get(f"/api/v1/executions/{exec_id}/export/xlsx")

    assert resp.status_code == 200
    workbook_path = tmp_path / "execution.xlsx"
    workbook_path.write_bytes(resp.content)
    workbook = openpyxl.load_workbook(workbook_path)
    summary = workbook["Summary"]
    labels = {row[0].value for row in summary.iter_rows()}

    assert any(label and "CET" in label or "CEST" in label for label in labels)


def test_xlsx_cases_export_header_includes_user(auth_client, project_id, case_ids, tmp_path):
    auth_client.put("/api/v1/users/me", json={"timezone": "Asia/Tokyo"})

    resp = auth_client.get(f"/api/v1/projects/{project_id}/cases/export/xlsx")

    assert resp.status_code == 200
    workbook_path = tmp_path / "cases.xlsx"
    workbook_path.write_bytes(resp.content)
    workbook = openpyxl.load_workbook(workbook_path)
    header = workbook.active["A2"].value
    assert header.startswith("Exported")
    assert "by u" in header
    assert "JST" in header
