"""Tests for the bug tracking API (P3.1)."""
import io

import openpyxl
import pytest
from fastapi.testclient import TestClient

from testjam.auth.security import hash_password
from testjam.main import app
from testjam.models.project import ProjectMember
from testjam.models.user import User
from tests.conftest import TestingSession


def _login(username: str, password: str = "pw") -> TestClient:
    client = TestClient(app)
    response = client.post(
        "/api/v1/auth/login", data={"username": username, "password": password},
    )
    client.headers["Authorization"] = f"Bearer {response.json()['access_token']}"
    return client


def _seed_member(project_id: int, username: str, role: str) -> int:
    with TestingSession() as db:
        user = User(
            username=username,
            email=f"{username}@x.com",
            hashed_password=hash_password("pw"),
            is_active=True,
        )
        db.add(user)
        db.commit()
        db.refresh(user)
        db.add(ProjectMember(project_id=project_id, user_id=user.id, role=role))
        db.commit()
        return user.id


@pytest.fixture
def project_id(auth_client):
    return auth_client.post("/api/v1/projects", json={"name": "Bug Project"}).json()["id"]


@pytest.fixture
def tester_client(project_id):
    _seed_member(project_id, "tester", "tester")
    return _login("tester")


@pytest.fixture
def viewer_client(project_id):
    _seed_member(project_id, "viewer", "viewer")
    return _login("viewer")


def _create_bug(client, project_id, **overrides):
    payload = {"title": "Login broken", "severity": "high"}
    payload.update(overrides)
    return client.post(f"/api/v1/projects/{project_id}/bugs", json=payload)


def test_create_bug_assigns_sequential_number(auth_client, project_id):
    first = _create_bug(auth_client, project_id, title="First").json()
    second = _create_bug(auth_client, project_id, title="Second").json()
    third = _create_bug(auth_client, project_id, title="Third").json()

    assert first["number"] == 1
    assert second["number"] == 2
    assert third["number"] == 3


def test_create_bug_records_initial_status_activity(auth_client, project_id):
    bug = _create_bug(auth_client, project_id).json()

    activity = auth_client.get(f"/api/v1/bugs/{bug['id']}/activity").json()

    assert len(activity) == 1
    assert activity[0]["field"] == "status"
    assert activity[0]["from_value"] is None
    assert activity[0]["to_value"] == "open"


def test_get_bug_by_number(auth_client, project_id):
    bug = _create_bug(auth_client, project_id).json()

    fetched = auth_client.get(f"/api/v1/projects/{project_id}/bugs/by-number/{bug['number']}")

    assert fetched.status_code == 200
    assert fetched.json()["id"] == bug["id"]


def test_list_bugs_filters_by_severity(auth_client, project_id):
    _create_bug(auth_client, project_id, severity="critical")
    _create_bug(auth_client, project_id, severity="low")

    crit = auth_client.get(f"/api/v1/projects/{project_id}/bugs?severity=critical").json()

    assert len(crit) == 1
    assert crit[0]["severity"] == "critical"


def test_list_bugs_filters_by_tag(auth_client, project_id):
    _create_bug(auth_client, project_id, tags=["crash"])
    _create_bug(auth_client, project_id, tags=["ui"])

    crashed = auth_client.get(f"/api/v1/projects/{project_id}/bugs?tag=crash").json()

    assert len(crashed) == 1
    assert "crash" in crashed[0]["tags"]


def test_update_bug_assigns_user(auth_client, project_id):
    bug = _create_bug(auth_client, project_id).json()
    me = auth_client.get("/api/v1/users/me").json()

    resp = auth_client.put(
        f"/api/v1/bugs/{bug['id']}", json={"assigned_to_id": me["id"]},
    )

    assert resp.status_code == 200
    assert resp.json()["assigned_to"]["id"] == me["id"]


def test_change_status_records_activity_and_resolved_at(auth_client, project_id):
    bug = _create_bug(auth_client, project_id).json()

    resolved = auth_client.post(
        f"/api/v1/bugs/{bug['id']}/status", json={"status": "resolved", "note": "Fixed in v2"},
    ).json()

    assert resolved["status"] == "resolved"
    assert resolved["resolved_at"] is not None
    activity = [
        row for row in auth_client.get(f"/api/v1/bugs/{bug['id']}/activity").json()
        if row["field"] == "status"
    ]
    assert [(row["from_value"], row["to_value"]) for row in activity] == [
        (None, "open"),
        ("open", "resolved"),
    ]
    assert activity[1]["note"] == "Fixed in v2"


def test_reopening_clears_resolved_at(auth_client, project_id):
    bug = _create_bug(auth_client, project_id).json()
    auth_client.post(f"/api/v1/bugs/{bug['id']}/status", json={"status": "closed"})

    reopened = auth_client.post(
        f"/api/v1/bugs/{bug['id']}/status", json={"status": "open"},
    ).json()

    assert reopened["resolved_at"] is None


def test_tester_can_create_bug_viewer_cannot(tester_client, viewer_client, project_id):
    created = tester_client.post(
        f"/api/v1/projects/{project_id}/bugs", json={"title": "From tester"},
    )
    blocked = viewer_client.post(
        f"/api/v1/projects/{project_id}/bugs", json={"title": "From viewer"},
    )

    assert created.status_code == 201
    assert blocked.status_code == 403


def test_owner_deletes_bug_tester_cannot(auth_client, tester_client, project_id):
    bug = _create_bug(auth_client, project_id).json()

    blocked = tester_client.delete(f"/api/v1/bugs/{bug['id']}")
    owner_resp = auth_client.delete(f"/api/v1/bugs/{bug['id']}")

    assert blocked.status_code == 403
    assert owner_resp.status_code == 204


def test_comment_round_trip(auth_client, project_id):
    bug = _create_bug(auth_client, project_id).json()

    created = auth_client.post(
        f"/api/v1/bugs/{bug['id']}/comments", json={"body": "Repro on staging"},
    )
    listing = auth_client.get(f"/api/v1/bugs/{bug['id']}/comments").json()

    assert created.status_code == 201
    assert created.json()["body"] == "Repro on staging"
    assert len(listing) == 1


def test_comment_author_can_edit_and_delete(tester_client, auth_client, project_id):
    bug = _create_bug(auth_client, project_id).json()
    comment = tester_client.post(
        f"/api/v1/bugs/{bug['id']}/comments", json={"body": "Original"},
    ).json()

    edited = tester_client.put(
        f"/api/v1/bugs/{bug['id']}/comments/{comment['id']}", json={"body": "Edited"},
    )
    deleted = tester_client.delete(
        f"/api/v1/bugs/{bug['id']}/comments/{comment['id']}",
    )

    assert edited.status_code == 200
    assert edited.json()["body"] == "Edited"
    assert deleted.status_code == 204


def test_environment_prefilled_from_execution(auth_client, project_id):
    execution_id = auth_client.post(
        f"/api/v1/projects/{project_id}/executions",
        json={"title": "Run", "type": "manual", "environment": "staging", "test_case_ids": []},
    ).json()["id"]

    bug = _create_bug(auth_client, project_id, execution_id=execution_id).json()

    assert bug["environment"] == "staging"


def test_explicit_environment_normalizes_to_slug(auth_client, project_id):
    bug = _create_bug(auth_client, project_id, environment="Production EU").json()
    assert bug["environment"] == "production-eu"


def test_html_report_renders_severity_groups(auth_client, project_id):
    _create_bug(auth_client, project_id, title="Crash on login", severity="critical")
    _create_bug(auth_client, project_id, title="Typo in footer", severity="low")

    response = auth_client.get(f"/api/v1/projects/{project_id}/bugs/report?format=html")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/html")
    body = response.text
    assert "Crash on login" in body
    assert "Typo in footer" in body
    assert "Critical" in body
    assert "Low" in body


def test_xlsx_report_returns_workbook(auth_client, project_id):
    _create_bug(auth_client, project_id, title="X")

    response = auth_client.get(f"/api/v1/projects/{project_id}/bugs/report?format=xlsx")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument"
    )
    assert response.content[:2] == b"PK"


def test_html_report_includes_comments_discussion(auth_client, project_id):
    bug = _create_bug(auth_client, project_id, title="Auth flake").json()
    auth_client.post(f"/api/v1/bugs/{bug['id']}/comments", json={"body": "Cannot reproduce on staging"})
    auth_client.post(f"/api/v1/bugs/{bug['id']}/comments", json={"body": "Repro steps in #42"})

    response = auth_client.get(f"/api/v1/projects/{project_id}/bugs/report?format=html")

    body = response.text
    assert "Discussion (2)" in body
    assert "Cannot reproduce on staging" in body
    assert "Repro steps in #42" in body


def test_xlsx_report_has_comments_sheet(auth_client, project_id):
    bug = _create_bug(auth_client, project_id, title="Crash").json()
    auth_client.post(f"/api/v1/bugs/{bug['id']}/comments", json={"body": "Happens on Safari only"})

    response = auth_client.get(f"/api/v1/projects/{project_id}/bugs/report?format=xlsx")
    workbook = openpyxl.load_workbook(io.BytesIO(response.content))

    assert "Comments" in workbook.sheetnames
    rows = list(workbook["Comments"].iter_rows(values_only=True))
    assert rows[0] == ("Bug #", "Bug title", "Author", rows[0][3], "Body")
    assert any("Happens on Safari only" in (row[4] or "") for row in rows[1:])


def test_update_bug_sets_updated_by(tester_client, auth_client, project_id):
    bug = _create_bug(auth_client, project_id).json()
    tester = tester_client.get("/api/v1/users/me").json()

    updated = tester_client.put(
        f"/api/v1/bugs/{bug['id']}", json={"description": "Edited by tester"},
    ).json()

    assert updated["updated_by"]["id"] == tester["id"]


def test_status_change_sets_updated_by(tester_client, auth_client, project_id):
    bug = _create_bug(auth_client, project_id).json()
    tester = tester_client.get("/api/v1/users/me").json()

    resolved = tester_client.post(
        f"/api/v1/bugs/{bug['id']}/status", json={"status": "resolved"},
    ).json()

    assert resolved["updated_by"]["id"] == tester["id"]


def test_get_bug_context_returns_execution_and_chain(auth_client, project_id):
    suite = auth_client.post(
        f"/api/v1/projects/{project_id}/suites", json={"name": "Auth"},
    ).json()
    case = auth_client.post(
        f"/api/v1/suites/{suite['id']}/cases",
        json={"name": "Login", "suite_id": suite["id"]},
    ).json()
    execution = auth_client.post(
        f"/api/v1/projects/{project_id}/executions",
        json={"title": "Sprint 1", "type": "manual", "test_case_ids": [case["id"]]},
    ).json()
    result = auth_client.get(f"/api/v1/executions/{execution['id']}/results").json()[0]
    bug = _create_bug(
        auth_client, project_id,
        result_id=result["id"], execution_id=execution["id"],
    ).json()

    context = auth_client.get(f"/api/v1/bugs/{bug['id']}/context").json()

    assert context["execution"] == {"id": execution["id"], "title": "Sprint 1"}
    assert context["case"]["name"] == "Login"
    assert [node["name"] for node in context["suite_path"]] == ["Auth"]


def test_link_crud_with_url(auth_client, project_id):
    bug = _create_bug(auth_client, project_id).json()

    created = auth_client.post(
        f"/api/v1/bugs/{bug['id']}/links",
        json={"url": "https://github.com/org/repo/pull/42", "label": "Fix PR"},
    )

    assert created.status_code == 201
    listing = auth_client.get(f"/api/v1/bugs/{bug['id']}/links").json()
    assert len(listing) == 1
    assert listing[0]["url"] == "https://github.com/org/repo/pull/42"
    assert listing[0]["label"] == "Fix PR"

    deleted = auth_client.delete(f"/api/v1/bugs/{bug['id']}/links/{created.json()['id']}")
    assert deleted.status_code == 204
    assert auth_client.get(f"/api/v1/bugs/{bug['id']}/links").json() == []


def test_link_to_execution_resolves_title(auth_client, project_id):
    execution = auth_client.post(
        f"/api/v1/projects/{project_id}/executions",
        json={"title": "Nightly", "type": "manual", "test_case_ids": []},
    ).json()
    bug = _create_bug(auth_client, project_id).json()

    created = auth_client.post(
        f"/api/v1/bugs/{bug['id']}/links",
        json={"execution_id": execution["id"]},
    ).json()

    assert created["execution_id"] == execution["id"]
    assert created["execution_title"] == "Nightly"


def test_link_to_target_bug_resolves_number(auth_client, project_id):
    bug_a = _create_bug(auth_client, project_id, title="Crash").json()
    bug_b = _create_bug(auth_client, project_id, title="Related").json()

    created = auth_client.post(
        f"/api/v1/bugs/{bug_a['id']}/links",
        json={"target_bug_id": bug_b["id"], "label": "relates to"},
    ).json()

    assert created["target_bug_number"] == bug_b["number"]
    assert created["target_bug_title"] == "Related"


def test_link_requires_at_least_one_target(auth_client, project_id):
    bug = _create_bug(auth_client, project_id).json()

    blocked = auth_client.post(
        f"/api/v1/bugs/{bug['id']}/links", json={"label": "nothing"},
    )

    assert blocked.status_code == 422


def test_link_kind_requires_target_bug(auth_client, project_id):
    bug = _create_bug(auth_client, project_id).json()

    blocked = auth_client.post(
        f"/api/v1/bugs/{bug['id']}/links",
        json={"kind": "blocks", "url": "https://example.com"},
    )

    assert blocked.status_code == 422


def test_blocks_creates_reciprocal_blocked_by(auth_client, project_id):
    blocker = _create_bug(auth_client, project_id, title="Blocker").json()
    blocked = _create_bug(auth_client, project_id, title="Blocked").json()

    auth_client.post(
        f"/api/v1/bugs/{blocker['id']}/links",
        json={"kind": "blocks", "target_bug_id": blocked["id"]},
    )

    reciprocal = auth_client.get(f"/api/v1/bugs/{blocked['id']}/links").json()
    assert len(reciprocal) == 1
    assert reciprocal[0]["kind"] == "blocked_by"
    assert reciprocal[0]["target_bug_id"] == blocker["id"]


def test_duplicate_of_is_symmetric(auth_client, project_id):
    a = _create_bug(auth_client, project_id, title="A").json()
    b = _create_bug(auth_client, project_id, title="B").json()

    auth_client.post(
        f"/api/v1/bugs/{a['id']}/links",
        json={"kind": "duplicate_of", "target_bug_id": b["id"]},
    )

    reciprocal = auth_client.get(f"/api/v1/bugs/{b['id']}/links").json()
    assert reciprocal[0]["kind"] == "duplicate_of"
    assert reciprocal[0]["target_bug_id"] == a["id"]


def test_deleting_link_removes_reciprocal(auth_client, project_id):
    a = _create_bug(auth_client, project_id, title="A").json()
    b = _create_bug(auth_client, project_id, title="B").json()
    created = auth_client.post(
        f"/api/v1/bugs/{a['id']}/links",
        json={"kind": "blocks", "target_bug_id": b["id"]},
    ).json()

    auth_client.delete(f"/api/v1/bugs/{a['id']}/links/{created['id']}")

    assert auth_client.get(f"/api/v1/bugs/{a['id']}/links").json() == []
    assert auth_client.get(f"/api/v1/bugs/{b['id']}/links").json() == []


def test_fixed_in_version_round_trip(auth_client, project_id):
    version = auth_client.post(
        f"/api/v1/projects/{project_id}/versions", json={"name": "v1.2"},
    ).json()
    bug = _create_bug(auth_client, project_id).json()

    updated = auth_client.put(
        f"/api/v1/bugs/{bug['id']}", json={"fixed_in_version_id": version["id"]},
    ).json()

    assert updated["fixed_in_version_id"] == version["id"]
    assert updated["fixed_in_version_name"] == "v1.2"


def test_link_kind_persists_in_response(auth_client, project_id):
    a = _create_bug(auth_client, project_id, title="A").json()
    b = _create_bug(auth_client, project_id, title="B").json()

    created = auth_client.post(
        f"/api/v1/bugs/{a['id']}/links",
        json={"kind": "relates_to", "target_bug_id": b["id"]},
    ).json()

    assert created["kind"] == "relates_to"


def _activity_by_field(client, bug_id, field):
    return [
        row for row in client.get(f"/api/v1/bugs/{bug_id}/activity").json()
        if row["field"] == field
    ]


def test_update_bug_records_field_diffs(auth_client, project_id):
    me = auth_client.get("/api/v1/users/me").json()
    bug = _create_bug(auth_client, project_id, title="Original", severity="medium").json()

    auth_client.put(
        f"/api/v1/bugs/{bug['id']}",
        json={
            "title": "Renamed",
            "severity": "critical",
            "tags": ["regression"],
            "assigned_to_id": me["id"],
        },
    )

    activity = auth_client.get(f"/api/v1/bugs/{bug['id']}/activity").json()
    by_field = {row["field"]: row for row in activity if row["field"] != "status"}

    assert by_field["title"]["from_value"] == "Original"
    assert by_field["title"]["to_value"] == "Renamed"
    assert by_field["severity"]["to_value"] == "critical"
    assert by_field["tags"]["to_value"] == '["regression"]'
    assert by_field["assigned_to"]["to_value"] == str(me["id"])


def test_update_bug_skips_unchanged_fields(auth_client, project_id):
    bug = _create_bug(auth_client, project_id, severity="medium").json()

    auth_client.put(f"/api/v1/bugs/{bug['id']}", json={"severity": "medium"})

    diffs = [row for row in auth_client.get(f"/api/v1/bugs/{bug['id']}/activity").json() if row["field"] != "status"]
    assert diffs == []


def test_link_add_and_delete_record_activity(auth_client, project_id):
    a = _create_bug(auth_client, project_id, title="A").json()
    b = _create_bug(auth_client, project_id, title="B").json()

    link = auth_client.post(
        f"/api/v1/bugs/{a['id']}/links",
        json={"kind": "relates_to", "target_bug_id": b["id"]},
    ).json()
    added = _activity_by_field(auth_client, a["id"], "link")
    assert len(added) == 1
    assert added[0]["from_value"] is None
    assert str(b["id"]) in added[0]["to_value"]

    auth_client.delete(f"/api/v1/bugs/{a['id']}/links/{link['id']}")

    rows = _activity_by_field(auth_client, a["id"], "link")
    assert len(rows) == 2
    assert rows[1]["from_value"] is not None
    assert rows[1]["to_value"] is None


def test_environment_change_records_assigned_field(auth_client, project_id):
    bug = _create_bug(auth_client, project_id).json()

    auth_client.put(f"/api/v1/bugs/{bug['id']}", json={"environment": "Staging"})

    env_rows = _activity_by_field(auth_client, bug["id"], "environment")
    assert len(env_rows) == 1
    assert env_rows[0]["from_value"] is None
    assert env_rows[0]["to_value"] == "staging"
